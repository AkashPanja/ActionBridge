from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
DONE_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
IN_PROGRESS_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
PENDING_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
BACKLOG_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, color="1E293B", size=14)
NORMAL_FONT = Font(size=10)
DONE_FONT = Font(size=10, color="166534")
PRIORITY_HIGH = Font(size=10, color="DC2626", bold=True)
PRIORITY_MEDIUM = Font(size=10, color="D97706")
PRIORITY_LOW = Font(size=10, color="6B7280")

thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

def style_header_row(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border
        if fill:
            cell.fill = fill

COLUMNS = ["ID", "Feature / Task", "Status", "Priority", "Sprint", "Effort", "Dependencies", "Notes"]
COL_WIDTHS = [6, 45, 14, 10, 10, 8, 20, 30]

# ─── SHEET 1: Complete Roadmap ───
ws1 = wb.active
ws1.title = "Sprint Roadmap"

ws1.cell(row=1, column=1, value="Action Bridge — Sprint Planning & Roadmap").font = TITLE_FONT
ws1.merge_cells("A1:H1")
ws1.row_dimensions[1].height = 30

ws1.cell(row=2, column=1, value="Last updated: June 2026 | Backend: FastAPI + SQLite/PostgreSQL | Frontend: React + Vite + Tailwind").font = Font(size=9, color="6B7280", italic=True)
ws1.merge_cells("A2:H2")

for i, (h, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
    ws1.cell(row=4, column=i, value=h)
    ws1.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws1, 4, 8)

# Data - all items from the full roadmap
items = [
    # Sprint 1-2: Core CRUD
    ("S1.1", "Backend CRUD: Projects, Document Types (JSON Schema)", "Done", "High", "Sprint 1", "5d", "", ""),
    ("S1.2", "Database setup + Alembic migrations", "Done", "High", "Sprint 1", "2d", "", ""),
    ("S1.3", "Docker Compose (PostgreSQL + app)", "Done", "Medium", "Sprint 1", "1d", "", ""),
    ("S1.4", "Frontend: Project + Document Type CRUD pages", "Done", "High", "Sprint 1", "4d", "", ""),
    ("S2.1", "Document lifecycle FSM (received→pending_review→approved/rejected)", "Done", "High", "Sprint 2", "3d", "", ""),
    ("S2.2", "Document submission with schema validation", "Done", "High", "Sprint 2", "2d", "", ""),
    ("S2.3", "Document inbox with filters + audit timeline", "Done", "High", "Sprint 2", "3d", "", ""),
    ("S2.4", "Schema-driven form in DocumentDetail + type coercion", "Done", "High", "Sprint 2", "4d", "", ""),
    ("S2.5", "Dark mode + glass-morphism design system", "Done", "Medium", "Sprint 2", "2d", "", ""),

    # Sprint 3: Auth
    ("S3.1", "JWT auth: User + ApiKey models", "Done", "High", "Sprint 3", "3d", "", ""),
    ("S3.2", "Permission-based RBAC (admin/reviewer/viewer)", "Done", "High", "Sprint 3", "2d", "", ""),
    ("S3.3", "Login/Register/Setup/Status endpoints + API key auth", "Done", "High", "Sprint 3", "2d", "", ""),
    ("S3.4", "Frontend: Login, Setup wizard, AuthContext", "Done", "High", "Sprint 3", "3d", "", ""),
    ("S3.5", "Setup wizard (replaced seed_admin, .setup-complete marker)", "Done", "Medium", "Sprint 3", "1d", "", ""),

    # Sprint 4: Schema Builder + Advanced Features
    ("S4.1", "Visual Schema Builder (add/remove/reorder fields, table support)", "Done", "High", "Sprint 4", "4d", "", ""),
    ("S4.2", "Advanced JSON Editor (undo/redo stacks, Ctrl+Z/Y)", "Done", "Medium", "Sprint 4", "2d", "", ""),
    ("S4.3", "Schema converters (flat fields ↔ table round-trip)", "Done", "Medium", "Sprint 4", "1d", "", ""),
    ("S4.4", "User Management (list/role/is_active toggle)", "Done", "High", "Sprint 4", "2d", "", ""),
    ("S4.5", "API Key management + Dashboard/Stats with charts", "Done", "High", "Sprint 4", "3d", "", ""),
    ("S4.6", "Bulk delete (projects, doc-types, documents)", "Done", "Medium", "Sprint 4", "2d", "", ""),
    ("S4.7", "Settings: Reset System with type-to-confirm", "Done", "Medium", "Sprint 4", "1d", "", ""),
    ("S4.8", "Rename app to 'Action Bridge'", "Done", "Low", "Sprint 4", "0.5d", "", ""),

    # Sprint 4b: Validation Rules Engine
    ("S4b.1", "Field-level validation rules (confidence_min, length, regex, range)", "Done", "High", "Sprint 4b", "4d", "S4.1", ""),
    ("S4b.2", "AND/OR pattern groups with negate flag", "Done", "High", "Sprint 4b", "2d", "", ""),
    ("S4b.3", "Built-in patterns (istext, isnumber, isemailid, isip, etc.)", "Done", "Medium", "Sprint 4b", "1d", "", ""),
    ("S4b.4", "RegexPattern model + CRUD at /api/v1/regex-patterns", "Done", "Medium", "Sprint 4b", "2d", "", ""),
    ("S4b.5", "ValidationRulesDialog: 3-source picker (builtin/saved/custom)", "Done", "High", "Sprint 4b", "3d", "", ""),
    ("S4b.6", "RegexBuilder component with live match highlighting", "Done", "Medium", "Sprint 4b", "2d", "", ""),
    ("S4b.7", "ValidationPatternsPage at /validation-patterns", "Done", "Medium", "Sprint 4b", "2d", "", ""),
    ("S4b.8", "Migration 004: regex_patterns table", "Done", "High", "Sprint 4b", "0.5d", "", ""),
    ("S4b.9", "confidence_min mandatory (default 95%, non-removable)", "Done", "High", "Sprint 4b", "0.5d", "", ""),
    ("S4b.10", "Auto-approve when all validations pass", "Done", "High", "Sprint 4b", "0.5d", "", ""),

    # Sprint 4c: Social Features
    ("S4c.1", "Email service: SMTP send with HTML/plain templates", "Done", "High", "Sprint 4c", "3d", "", ""),
    ("S4c.2", "Models: ProjectMembership, Notification, DocTypeSubscription, Comment, Attachment", "Done", "High", "Sprint 4c", "2d", "", ""),
    ("S4c.3", "Project model: created_by + visibility columns", "Done", "High", "Sprint 4c", "0.5d", "", ""),
    ("S4c.4", "Settings router: Company/SMTP/Password Policy CRUD + logo upload", "Done", "High", "Sprint 4c", "2d", "", ""),
    ("S4c.5", "Notifications router (list/unread/mark-read/mark-all-read)", "Done", "Medium", "Sprint 4c", "1d", "", ""),
    ("S4c.6", "Comments router (threaded create/edit/delete)", "Done", "Medium", "Sprint 4c", "1.5d", "", ""),
    ("S4c.7", "Attachments router (upload/list/delete with mime+size filter)", "Done", "Medium", "Sprint 4c", "1.5d", "", ""),
    ("S4c.8", "Invitations router (list-pending/accept/decline)", "Done", "Medium", "Sprint 4c", "1d", "", ""),
    ("S4c.9", "Subscriptions router (list/upsert/delete)", "Done", "Low", "Sprint 4c", "0.5d", "", ""),
    ("S4c.10", "Auth: user create/delete, password change, password reset flow", "Done", "High", "Sprint 4c", "2d", "", ""),
    ("S4c.11", "Project service: memberships (invite/accept/remove/role)", "Done", "High", "Sprint 4c", "2d", "", ""),
    ("S4c.12", "Backend main: register all routers + static uploads mount", "Done", "High", "Sprint 4c", "0.5d", "", ""),
    ("S4c.13", "Migration 005: all new tables + projects FKs (SQLite batch mode)", "Done", "High", "Sprint 4c", "1d", "", ""),

    # Sprint 4d: Frontend Social
    ("S4d.1", "SettingsPage: tabbed UI (Company/SMTP/Password/Danger)", "Done", "High", "Sprint 4d", "2d", "S4c.4", ""),
    ("S4d.2", "TopBar: notification bell dropdown with mark-read + password change modal", "Done", "High", "Sprint 4d", "2d", "S4c.5", ""),
    ("S4d.3", "DocumentDetail: threaded comments section + attachments upload", "Done", "High", "Sprint 4d", "3d", "S4c.6+S4c.7", ""),
    ("S4d.4", "UserManagement: create user form + delete button (self-protected)", "Done", "High", "Sprint 4d", "1.5d", "S4c.10", ""),
    ("S4d.5", "MembersPage: invite/remove/change-role with user selector", "Done", "High", "Sprint 4d", "2d", "S4c.11", ""),
    ("S4d.6", "ProjectDetail: Members tab added", "Done", "Medium", "Sprint 4d", "0.5d", "", ""),
    ("S4d.7", "DocumentTypeList: bell toggle subscription per card", "Done", "Medium", "Sprint 4d", "1d", "S4c.9", ""),
    ("S4d.8", "ForgotPasswordPage + ResetPasswordPage (2-step flow)", "Done", "Medium", "Sprint 4d", "1.5d", "S4c.10", ""),
    ("S4d.9", "Settings SMTP: Send Test Email button", "Done", "Low", "Sprint 4d", "0.5d", "", ""),

    # Sprint 5: Dashboard & Polish
    ("S5.1", "Dashboard charts: daily volume line chart (last 7 days)", "In Progress", "High", "Sprint 5", "2d", "S4d.6", "Use recharts; fetch stats from /projects/{id}/stats"),
    ("S5.2", "Dashboard charts: status distribution pie/bar chart", "In Progress", "High", "Sprint 5", "1.5d", "", ""),
    ("S5.3", "Dashboard: recent activity feed (last 5 audit events)", "In Progress", "Medium", "Sprint 5", "1.5d", "", ""),
    ("S5.4", "Invitations pending badge in sidebar", "Pending", "Medium", "Sprint 5", "1d", "S4c.8", "Fetch GET /api/v1/invitations count"),
    ("S5.5", "Loading/empty/error states for comments, attachments, members", "Pending", "Medium", "Sprint 5", "2d", "", "Refine existing components"),
    ("S5.6", "Error boundaries for all pages", "Pending", "Low", "Sprint 5", "1d", "", "React error boundary pattern"),
    ("S5.7", "Document detail: download attachment link (a href to /uploads/...)", "Pending", "Low", "Sprint 5", "0.5d", "S4c.7", "Make attachment filename clickable"),

    # Sprint 6: Testing
    ("S6.1", "Backend: pytest setup + conftest (test DB, auth fixtures)", "Pending", "High", "Sprint 6", "2d", "", ""),
    ("S6.2", "Backend: test auth endpoints (login, register, setup, me, users)", "Pending", "High", "Sprint 6", "2d", "", ""),
    ("S6.3", "Backend: test project CRUD + member invite/accept/remove", "Pending", "High", "Sprint 6", "2d", "", ""),
    ("S6.4", "Backend: test document submission, validation rules, FSM transitions", "Pending", "High", "Sprint 6", "3d", "", ""),
    ("S6.5", "Backend: test comments, attachments, notifications, subscriptions", "Pending", "Medium", "Sprint 6", "2d", "", ""),
    ("S6.6", "Backend: test regex-patterns CRUD + pattern resolution in validation", "Pending", "Medium", "Sprint 6", "1.5d", "", ""),
    ("S6.7", "Frontend: vitest + testing-library setup", "Pending", "Medium", "Sprint 6", "1d", "", ""),
    ("S6.8", "Frontend: test schema builder, validation rules dialog, regex builder", "Pending", "Medium", "Sprint 6", "3d", "", ""),
    ("S6.9", "Frontend: test auth flow (login, setup, password reset)", "Pending", "Medium", "Sprint 6", "2d", "", ""),

    # Sprint 7: Deployment & CI/CD
    ("S7.1", "Docker Compose for production (backend + frontend + nginx)", "Pending", "High", "Sprint 7", "2d", "", ""),
    ("S7.2", "GitHub Actions CI: lint + typecheck + test on PR", "Pending", "High", "Sprint 7", "1.5d", "Sprint 6", ""),
    ("S7.3", "GitHub Actions CD: deploy on merge to main", "Pending", "Medium", "Sprint 7", "1.5d", "", ""),
    ("S7.4", "Environment variable documentation (.env.example full)", "Pending", "Low", "Sprint 7", "0.5d", "", ""),
    ("S7.5", "PostgreSQL migration production check", "Pending", "High", "Sprint 7", "1d", "", "Test Alembic migrations against fresh PostgreSQL"),

    # Backlog / Future
    ("B.1", "Webhook support: notify external systems on document events", "Backlog", "Low", "Future", "3d", "", "POST to configured webhook URL on status change"),
    ("B.2", "Export documents to CSV/Excel", "Backlog", "Low", "Future", "2d", "", ""),
    ("B.3", "Multi-page document support (batch submission)", "Backlog", "Low", "Future", "4d", "", "Submit multiple documents at once"),
    ("B.4", "Activity log page (global audit events, not per-document)", "Backlog", "Low", "Future", "2d", "", ""),
    ("B.5", "Dark mode persistence (localStorage/user preference)", "Backlog", "Low", "Future", "0.5d", "", ""),
    ("B.6", "Internationalization (i18n) support", "Backlog", "Low", "Future", "5d", "", ""),
]

for r, item in enumerate(items, 5):
    for c, val in enumerate(item, 1):
        ws1.cell(row=r, column=c, value=val)

    # Color code by status
    if item[2] == "Done":
        fill = DONE_FILL
    elif item[2] == "In Progress":
        fill = IN_PROGRESS_FILL
    elif item[2] == "Backlog":
        fill = BACKLOG_FILL
    else:
        fill = PENDING_FILL

    style_data_row(ws1, r, 8, fill)

    # Priority coloring
    prio_cell = ws1.cell(row=r, column=4)
    if item[3] == "High":
        prio_cell.font = PRIORITY_HIGH
    elif item[3] == "Medium":
        prio_cell.font = PRIORITY_MEDIUM
    else:
        prio_cell.font = PRIORITY_LOW

# Freeze pane
ws1.freeze_panes = "A5"
ws1.auto_filter.ref = f"A4:H{4 + len(items)}"

for r in range(5, 5 + len(items)):
    ws1.row_dimensions[r].height = 22

# ─── SHEET 2: Sprint 5 Detail ───
ws2 = wb.create_sheet("Sprint 5 — Dashboard & Polish")
ws2.cell(row=1, column=1, value="Sprint 5: Dashboard & Polish").font = TITLE_FONT
ws2.merge_cells("A1:G1")
ws2.cell(row=2, column=1, value="Goal: Complete the project dashboard with charts, polish new features, handle edge cases").font = Font(size=9, color="6B7280", italic=True)
ws2.merge_cells("A2:G2")

S5_COLS = ["Step", "Task", "Status", "Effort", "Files to modify", "Acceptance Criteria", "Notes"]
S5_WIDTHS = [5, 40, 12, 8, 35, 35, 30]

for i, (h, w) in enumerate(zip(S5_COLS, S5_WIDTHS), 1):
    ws2.cell(row=4, column=i, value=h)
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws2, 4, 7)

s5_items = [
    ("1", "Install recharts package", "Pending", "0.1d", "frontend/package.json", "npm install recharts", ""),
    ("2", "Dashboard: daily volume line chart", "Pending", "2d",
     "frontend/src/pages/ProjectDashboard.tsx",
     "Line chart shows 7-day document volume from /projects/{id}/stats.daily_volume; dates on X, count on Y; different colored lines per status",
     "Use ResponsiveContainer, XAxis/YAxis, Tooltip, Legend, CartesianGrid, Line"),
    ("3", "Dashboard: status distribution pie chart", "Pending", "1.5d",
     "frontend/src/pages/ProjectDashboard.tsx",
     "Pie chart shows breakdown from stats.status_breakdown; labels with count + %, tooltip, color map per status (emerald=approved, amber=pending, red=rejected, gray=received)",
     "Use PieChart, Pie, Cell, Tooltip"),
    ("4", "Dashboard: recent activity feed", "Pending", "1.5d",
     "frontend/src/pages/ProjectDashboard.tsx",
     "Shows latest 5 audit events across all documents in project; displays action, document type, timestamp, actor; clickable to navigate to document detail",
     "Add new endpoint or reuse existing audit data; show skeleton while loading"),
    ("5", "Invitations pending badge in sidebar", "Pending", "1d",
     "frontend/src/components/layout/Sidebar.tsx, frontend/src/contexts/AuthContext.tsx",
     "Sidebar shows small red badge with pending invitation count next to bell or user section; auto-fetches every 30s",
     "Fetch GET /api/v1/invitations, count items with status=pending"),
    ("6", "Loading/empty/error states polish", "Pending", "2d",
     "Various: DocumentDetail, MembersPage, DocumentTypeList, TopBar",
     "All new components show Skeleton while loading, empty state message when no data, error banner when API fails",
     "Already partially done; audit and fix any missing states"),
    ("7", "Attachment download link", "Pending", "0.5d",
     "frontend/src/pages/documents/DocumentDetail.tsx",
     "Attachment file name is a clickable link that opens /uploads/{file_path} in new tab",
     "Right now it's just a display; make it a functional <a> tag"),
    ("8", "Error boundary component", "Pending", "1d",
     "frontend/src/components/shared/ErrorBoundary.tsx",
     "Wrap each route page in ErrorBoundary; shows friendly error UI with retry button",
     "Standard React error boundary + fallback UI"),
    ("9", "Final verification: tsc + build + backend start", "Pending", "0.5d",
     "", "All TypeScript compiles, frontend builds, backend starts without errors",
     ""),
]

for r, item in enumerate(s5_items, 5):
    for c, val in enumerate(item, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, 7, PENDING_FILL if item[2] == "Pending" else DONE_FILL)
    ws2.row_dimensions[r].height = 50

ws2.freeze_panes = "A5"

# ─── SHEET 3: Sprint 6 Detail ───
ws3 = wb.create_sheet("Sprint 6 — Testing")
ws3.cell(row=1, column=1, value="Sprint 6: Testing").font = TITLE_FONT
ws3.merge_cells("A1:G1")
ws3.cell(row=2, column=1, value="Goal: Add comprehensive test coverage for backend API endpoints and frontend components").font = Font(size=9, color="6B7280", italic=True)
ws3.merge_cells("A2:G2")

for i, (h, w) in enumerate(zip(S5_COLS, S5_WIDTHS), 1):
    ws3.cell(row=4, column=i, value=h)
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws3, 4, 7)

s6_items = [
    ("1", "Backend: pytest + httpx + test DB setup", "Pending", "2d",
     "backend/conftest.py, backend/tests/conftest.py",
     "async test client with SQLite test DB, fixture for admin user, fixture for auth headers",
     "Use pytest-asyncio, httpx.AsyncClient, separate test DB file"),
    ("2", "Backend: test auth endpoints", "Pending", "2d",
     "backend/tests/test_auth.py",
     "Tests for: setup, register, login, me, users CRUD, password change, password reset, API key CRUD",
     ""),
    ("3", "Backend: test project + membership", "Pending", "2d",
     "backend/tests/test_projects.py",
     "Tests for: create/read/update/delete project, list with visibility filter, invite/accept/decline/remove/role-change",
     ""),
    ("4", "Backend: test document + validation", "Pending", "3d",
     "backend/tests/test_documents.py",
     "Tests for: submit document, schema validation, validation rules (confidence, patterns, AND/OR groups), FSM transitions, audit events, bulk ops",
     ""),
    ("5", "Backend: test new routers (comments, attachments, notifications, subscriptions)", "Pending", "2d",
     "backend/tests/test_social.py",
     "Tests for: threaded comments CRUD, attachment upload/list/delete, notification mark-read, subscription upsert/delete",
     ""),
    ("6", "Frontend: vitest + testing-library setup", "Pending", "1d",
     "frontend/vitest.config.ts, frontend/src/test-setup.ts",
     "vitest configured with jsdom, testing-library render helpers, mock for API calls",
     ""),
    ("7", "Frontend: test SchemaBuilder + ValidationRulesDialog", "Pending", "3d",
     "frontend/src/tests/",
     "Tests for: field add/remove/reorder, type change, table column editor, validation rules picker, confidence slider",
     "Use MSW or manual fetch mocks"),
    ("8", "Frontend: test auth flow", "Pending", "2d",
     "frontend/src/tests/",
     "Tests for: login, setup wizard, forgot/reset password, auth context behavior",
     ""),
]

for r, item in enumerate(s6_items, 5):
    for c, val in enumerate(item, 1):
        ws3.cell(row=r, column=c, value=val)
    style_data_row(ws3, r, 7, PENDING_FILL)
    ws3.row_dimensions[r].height = 55

ws3.freeze_panes = "A5"

# ─── SHEET 4: Sprint 7 Detail ───
ws4 = wb.create_sheet("Sprint 7 — Deployment and CI-CD")
ws4.cell(row=1, column=1, value="Sprint 7: Deployment & CI/CD").font = TITLE_FONT
ws4.merge_cells("A1:G1")
ws4.cell(row=2, column=1, value="Goal: Production-ready deployment with CI/CD pipeline").font = Font(size=9, color="6B7280", italic=True)
ws4.merge_cells("A2:G2")

for i, (h, w) in enumerate(zip(S5_COLS, S5_WIDTHS), 1):
    ws4.cell(row=4, column=i, value=h)
    ws4.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws4, 4, 7)

s7_items = [
    ("1", "Production Docker Compose", "Pending", "2d",
     "docker-compose.prod.yml, frontend/Dockerfile, backend/Dockerfile, nginx/",
     "Multi-service compose: backend (gunicorn + uvicorn workers), frontend (nginx serving static build), postgres; env vars via .env.prod",
     "Use multi-stage build for frontend"),
    ("2", "GitHub Actions CI pipeline", "Pending", "1.5d",
     ".github/workflows/ci.yml",
     "On PR: run backend tests (pytest), run frontend typecheck + build, lint (ruff for Python, ESLint for TS)",
     ""),
    ("3", "GitHub Actions CD pipeline", "Pending", "1.5d",
     ".github/workflows/cd.yml",
     "On merge to main: build + push Docker images to registry, deploy to VPS (SSH + docker-compose pull/up)",
     ""),
    ("4", "Env docs + .env.example cleanup", "Pending", "0.5d",
     "backend/.env.example",
     "Full documentation of all env vars: SECRET_KEY, DATABASE_URL, SMTP_*, UPLOAD_DIR",
     ""),
    ("5", "PostgreSQL migration verification", "Pending", "1d",
     "backend/alembic/",
     "Run all migrations against fresh PostgreSQL DB; fix any dialect-specific SQL issues (JSON, timestamps, etc.)",
     "SQLite has some leniencies that Postgres does not"),
]

for r, item in enumerate(s7_items, 5):
    for c, val in enumerate(item, 1):
        ws4.cell(row=r, column=c, value=val)
    style_data_row(ws4, r, 7, PENDING_FILL)
    ws4.row_dimensions[r].height = 55

ws4.freeze_panes = "A5"

# ─── SHEET 5: Backlog ───
ws5 = wb.create_sheet("Backlog & Future")
ws5.cell(row=1, column=1, value="Backlog & Future Features").font = TITLE_FONT
ws5.merge_cells("A1:F1")

BACKLOG_COLS = ["ID", "Feature", "Effort", "Priority", "Dependencies", "Notes"]
BACKLOG_WIDTHS = [5, 40, 8, 10, 25, 35]

for i, (h, w) in enumerate(zip(BACKLOG_COLS, BACKLOG_WIDTHS), 1):
    ws5.cell(row=3, column=i, value=h)
    ws5.column_dimensions[get_column_letter(i)].width = w
style_header_row(ws5, 3, 6, fill=SUBHEADER_FILL)

backlog_items = [
    ("B.1", "Webhook support: notify external systems on document events", "3d", "Low", "Email service", "POST to configured webhook URL on status change"),
    ("B.2", "Export documents to CSV/Excel", "2d", "Low", "", "Download all documents in a project as a spreadsheet"),
    ("B.3", "Multi-page document batch submission", "4d", "Low", "", "Submit multiple documents in one API call / form"),
    ("B.4", "Global activity log page", "2d", "Low", "", "View all audit events across all projects"),
    ("B.5", "Dark mode persistence (localStorage)", "0.5d", "Low", "", "Remember dark mode preference across sessions"),
    ("B.6", "i18n / localization support", "5d", "Low", "", "React-Intl or react-i18next setup with English as base"),
    ("B.7", "Document template cloning (copy doc type to another project)", "1.5d", "Low", "Sprint 4b", "Clone entire schema + validation rules"),
    ("B.8", "Comment @mentions + in-app notification", "2d", "Low", "Comments + Notifications", "@username triggers notification to that user"),
]

for r, item in enumerate(backlog_items, 4):
    for c, val in enumerate(item, 1):
        ws5.cell(row=r, column=c, value=val)
    style_data_row(ws5, r, 6, fill=BACKLOG_FILL)
    ws5.row_dimensions[r].height = 30

ws5.freeze_panes = "A4"

# Save
output_path = "ActionBridge_Sprint_Planning.xlsx"
wb.save(output_path)
print(f"[OK] Sprint planning Excel saved to: {output_path}")
